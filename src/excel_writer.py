"""
Модуль создания итоговой Excel таблицы из JSON данных.
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# === Папка для Excel (корень проекта, а не src) ===
# Поднимаемся на уровень выше из папки src
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_EXCEL_DIR = os.path.join(BASE_DIR, "output_excel")
os.makedirs(OUTPUT_EXCEL_DIR, exist_ok=True)

# Колонки для Excel
COLUMNS = [
    "Поз.",
    "Наименование и техническая характеристика",
    "Тип, марка, обозначение документа, опросного листа",
    "Код продукции",
    "Поставщик",
    "Единица измерения",
    "Количество",
    "Масса 1 ед.,кг",
    "Примечание",
    "Источник"
]

# Порядок ключей JSON -> Excel
KEYS_ORDER = ["pos", "name", "type_mark", "product_code", "supplier",
              "unit", "quantity", "mass", "note", "source"]

# Ширины столбцов
COLUMN_WIDTHS = {
    1: 8,    # Поз.
    2: 45,   # Наименование
    3: 35,   # Тип, марка
    4: 15,   # Код продукции
    5: 20,   # Поставщик
    6: 12,   # Единица измерения
    7: 12,   # Количество
    8: 14,   # Масса
    9: 20,   # Примечание
    10: 20   # Источник
}


def create_excel_from_json(data: list, output_filename: str = None) -> str:
    """Создаёт отформатированный Excel файл из данных спецификации."""
    os.makedirs(OUTPUT_EXCEL_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    # --- Стили ---
    header_font_white = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="Times New Roman", size=10)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # --- Заголовок документа ---
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "СПЕЦИФИКАЦИЯ ОБОРУДОВАНИЯ, ИЗДЕЛИЙ И МАТЕРИАЛОВ"
    title_cell.font = Font(name="Times New Roman", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Заголовки столбцов ---
    header_row = 3
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_alignment
    
    ws.row_dimensions[header_row].height = 45

    # --- Ширины столбцов ---
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Данные ---
    center_columns = {1, 6, 7, 8}

    for row_idx, item in enumerate(data, header_row + 1):
        for col_idx, key in enumerate(KEYS_ORDER, 1):
            value = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_alignment if col_idx in center_columns else data_alignment
        ws.row_dimensions[row_idx].height = 30

    # --- Итоговая строка ---
    total_row = header_row + len(data) + 1
    ws.merge_cells(f"A{total_row}:E{total_row}")
    total_cell = ws.cell(row=total_row, column=1, value=f"Итого позиций: {len(data)}")
    total_cell.font = Font(name="Times New Roman", size=11, bold=True)
    total_cell.alignment = Alignment(horizontal="right")

    # --- Дата создания ---
    date_row = total_row + 2
    ws.cell(row=date_row, column=1, value=f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row=date_row, column=1).font = Font(name="Times New Roman", size=9, italic=True)

    # --- Сохранение ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if output_filename:
        excel_filename = f"{output_filename}.xlsx"
    else:
        excel_filename = f"specification_{timestamp}.xlsx"

    excel_path = os.path.join(OUTPUT_EXCEL_DIR, excel_filename)
    wb.save(excel_path)

    logger.info(f"Excel создан: {excel_path} ({len(data)} записей)")
    return excel_path